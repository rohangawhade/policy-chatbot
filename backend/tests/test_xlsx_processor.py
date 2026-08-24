from pathlib import Path

import openpyxl
import pytest

from adapters.document_processors.xlsx_processor import XLSXProcessor
from core.ports.document_processor_port import DocumentProcessorPort


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "sample.xlsx"
    workbook = openpyxl.Workbook()
    coverage_sheet = workbook.active
    assert coverage_sheet is not None
    coverage_sheet.title = "Coverage"
    coverage_sheet.append(["Plan", "Deductible"])
    coverage_sheet.append(["Dental PPO", 500])

    rates_sheet = workbook.create_sheet("Rates")
    rates_sheet.append(["Tier", "Monthly Cost"])
    rates_sheet.append(["Employee Only", 25.5])

    workbook.save(path)
    return path


def test_is_a_document_processor_port() -> None:
    assert isinstance(XLSXProcessor(), DocumentProcessorPort)


def test_extract_text_includes_every_sheets_rows(sample_xlsx: Path) -> None:
    text = XLSXProcessor().extract_text(str(sample_xlsx))

    assert "# Coverage" in text
    assert "Plan | Deductible" in text
    assert "Dental PPO | 500" in text
    assert "# Rates" in text
    assert "Tier | Monthly Cost" in text
    assert "Employee Only | 25.5" in text


def test_extract_text_skips_fully_empty_rows(tmp_path: Path) -> None:
    path = tmp_path / "with_blanks.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["first"])
    sheet.append([None])
    sheet.append(["second"])
    workbook.save(path)

    text = XLSXProcessor().extract_text(str(path))

    assert text == "# Sheet\nfirst\nsecond"


def test_extract_metadata_returns_sheet_names_and_count(sample_xlsx: Path) -> None:
    metadata = XLSXProcessor().extract_metadata(str(sample_xlsx))

    assert metadata["sheet_names"] == ["Coverage", "Rates"]
    assert metadata["sheet_count"] == 2
