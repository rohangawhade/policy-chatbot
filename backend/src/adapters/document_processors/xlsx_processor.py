"""XLSX document processor: openpyxl for cell-level, sheet-aware
extraction (files/plan.md Step 3.6)."""

from typing import Any

import openpyxl

from core.ports.document_processor_port import DocumentProcessorPort


class XLSXProcessor(DocumentProcessorPort):
    def extract_text(self, file_path: str) -> str:
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            sheet_blocks = []
            for sheet in workbook.worksheets:
                row_lines = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(value) for value in row if value is not None]
                    if cells:
                        row_lines.append(" | ".join(cells))
                if row_lines:
                    sheet_blocks.append(f"# {sheet.title}\n" + "\n".join(row_lines))
            return "\n\n".join(sheet_blocks)
        finally:
            workbook.close()

    def extract_metadata(self, file_path: str) -> dict[str, Any]:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        try:
            return {
                "sheet_names": list(workbook.sheetnames),
                "sheet_count": len(workbook.sheetnames),
            }
        finally:
            workbook.close()
