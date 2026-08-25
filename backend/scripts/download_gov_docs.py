"""Downloads real, publicly-available benefits documents from four
government sources for the RAG corpus (files/plan.md Step 11.1):

- OPM.gov: every distinct FEHB carrier plan brochure, discovered in
  bulk from OPM's own published "FEHB Plan Key" spreadsheet (a public
  use file mapping each plan to its brochure number) rather than a
  hand-maintained URL list -- 58 unique, well-formed brochure numbers
  as of the 2026 plan year (one further row is malformed in OPM's own
  data and is skipped, see `_discover_opm_brochures`), which alone
  covers most of this step's 50-100 document target.
- CMS.gov / Medicare.gov: SBC (Summary of Benefits and Coverage)
  templates/samples and Medicare & Medicaid summary documents.
- DOL.gov (EBSA): ERISA compliance guides.

**Interpretation of "healthcare.gov -> SBC templates" (plan.md)**:
healthcare.gov's own domain doesn't host the SBC template/sample PDFs
directly -- its own SBC page links out to CMS.gov/CCIIO, which is where
the actual files live (confirmed by fetching
healthcare.gov/health-care-law-protections/summary-of-benefits-and-coverage/
during research for this script). Filed under CMS here rather than
duplicating the same files under a "healthcare_gov" folder that would
just re-download identical content from the same origin.

Every request sends a descriptive, honest User-Agent identifying this
as an automated downloader (not a spoofed browser string) -- several of
these sites 403 a bare/default User-Agent but happily serve a named one;
that's ordinary bot etiquette, not evasion of any real access control.

Idempotent by design: a manifest (`data/gov_pdfs/manifest.json`) records
every file already downloaded (source, url, sha256, size, timestamp);
re-running the script only fetches what's missing unless --force is
passed. Safe to interrupt and re-run.

Usage:
    python scripts/download_gov_docs.py [--dry-run] [--force]
        [--source {opm,cms,dol,all}] [--limit N]

    --dry-run   List what would be downloaded without downloading it.
    --force     Re-download even files the manifest already has.
    --source    Restrict to one source (default: all).
    --limit N   Cap the number of files downloaded per source (testing).
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import io
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

import httpx
import openpyxl
import structlog
from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential

logger = structlog.get_logger(__name__)

_REPO_ROOT = Path(__file__).resolve().parents[2]
_OUTPUT_ROOT = _REPO_ROOT / "data" / "gov_pdfs"
_MANIFEST_PATH = _OUTPUT_ROOT / "manifest.json"

_USER_AGENT = "PolicyPalDocBot/1.0 (+educational project; contact: gawhaderohan.rg@gmail.com)"

# The filename embeds the publication date and changes each plan year --
# re-verify/update this URL when re-running for a new FEHB plan year by
# browsing https://www.opm.gov/healthcare-insurance/healthcare/
# transparency-in-healthcare/public-use-files/ for the current link.
_OPM_PLAN_KEY_URL = (
    "https://www.opm.gov/healthcare-insurance/healthcare/transparency-in-healthcare/"
    "public-use-files/2026/fehb/2026-fehb-plan-key-11202025.xlsx"
)
_OPM_BROCHURE_URL_TEMPLATE = (
    "https://www.opm.gov/healthcare-insurance/healthcare/plan-information/"
    "plans/pdf/2026/brochures/{code}.pdf"
)

# Curated, individually-verified real documents (unlike OPM, CMS/DOL have
# no single public index to bulk-discover these from -- each URL below
# was confirmed to return 200 with a real PDF during this step's research).
_CMS_DOCS: list[tuple[str, str, str]] = [
    # (subfolder, filename, url)
    (
        "sbc_templates",
        "sbc-template.pdf",
        "https://www.cms.gov/CCIIO/Resources/Regulations-and-Guidance/Downloads/SBC-Template.pdf",
    ),
    (
        "sbc_templates",
        "sbc-sample-completed.pdf",
        "https://www.cms.gov/cciio/resources/forms-reports-and-other-resources/downloads/"
        "sbc-sample-completed-mm-508-fixed-4-12-16.pdf",
    ),
    (
        "medicare_medicaid_summaries",
        "brief-summaries-medicare-medicaid.pdf",
        "https://www.cms.gov/files/document/"
        "brief-summaries-medicare-medicaid-november-15-2024.pdf",
    ),
    (
        "medicare_medicaid_summaries",
        "medicare-and-you-handbook.pdf",
        "https://www.medicare.gov/publications/10050-medicare-and-you.pdf",
    ),
    (
        "medicare_medicaid_summaries",
        "medicaid-chip-fast-facts.pdf",
        "https://www.cms.gov/marketplace/technical-assistance-resources/"
        "fast-facts-medicaid-chip.pdf",
    ),
    (
        "medicare_medicaid_summaries",
        "medicaid-program-overview-and-policy.pdf",
        "https://www.cms.gov/files/document/chapter-1-program-overview-and-policy.pdf",
    ),
]

_DOL_DOCS: list[tuple[str, str, str]] = [
    (
        "erisa_guides",
        "self-compliance-tool-erisa-part-7.pdf",
        "https://www.dol.gov/sites/dolgov/files/EBSA/about-ebsa/our-activities/"
        "resource-center/publications/compliance-self-assessment.pdf",
    ),
    (
        "erisa_guides",
        "employees-guide-to-cobra.pdf",
        "https://www.dol.gov/sites/dolgov/files/EBSA/about-ebsa/our-activities/"
        "resource-center/publications/an-employees-guide-to-health-benefits-under-cobra.pdf",
    ),
    (
        "erisa_guides",
        "fiduciary-responsibilities-group-health-plan.pdf",
        "https://www.dol.gov/sites/dolgov/files/ebsa/about-ebsa/our-activities/"
        "resource-center/publications/"
        "understanding-your-fiduciary-responsibilities-under-a-group-health-plan.pdf",
    ),
    (
        "erisa_guides",
        "reporting-and-disclosure-guide.pdf",
        "https://www.dol.gov/sites/dolgov/files/EBSA/about-ebsa/our-activities/"
        "resource-center/publications/reporting-and-disclosure-guide-for-employee-benefit-plans.pdf",
    ),
    (
        "erisa_guides",
        "mewa-guide.pdf",
        "https://www.dol.gov/sites/dolgov/files/EBSA/about-ebsa/our-activities/"
        "resource-center/publications/mewa-under-erisa-a-guide-to-federal-and-state-regulation.pdf",
    ),
]

_SOURCES = Literal["opm", "cms", "dol"]


@dataclass(frozen=True, kw_only=True)
class DownloadTarget:
    source: _SOURCES
    subfolder: str
    filename: str
    url: str


@dataclass
class Manifest:
    entries: dict[str, dict[str, str | int]] = field(default_factory=dict)

    @classmethod
    def load(cls) -> Manifest:
        if not _MANIFEST_PATH.exists():
            return cls()
        return cls(entries=json.loads(_MANIFEST_PATH.read_text(encoding="utf-8")))

    def save(self) -> None:
        _MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
        _MANIFEST_PATH.write_text(json.dumps(self.entries, indent=2, sort_keys=True) + "\n")

    def has(self, relative_path: str) -> bool:
        return relative_path in self.entries

    def record(self, relative_path: str, *, url: str, content: bytes) -> None:
        self.entries[relative_path] = {
            "url": url,
            "sha256": hashlib.sha256(content).hexdigest(),
            "size_bytes": len(content),
            "downloaded_at": datetime.now(UTC).isoformat(),
        }


@retry(
    retry=retry_if_exception_type(httpx.HTTPError),
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    reraise=True,
)
async def _fetch(client: httpx.AsyncClient, url: str) -> bytes:
    response = await client.get(url, follow_redirects=True, timeout=30.0)
    response.raise_for_status()
    return response.content


_BROCHURE_CODE_RE = re.compile(r"(\d{2}-\d{3,4})")


async def _discover_opm_brochures(client: httpx.AsyncClient) -> list[DownloadTarget]:
    """Downloads OPM's own FEHB Plan Key spreadsheet and turns every
    distinct "Brochure Number" column value ("RI 73-860") into a real
    brochure PDF URL ("73-860.pdf") -- verified during this step's
    research that this transform matches OPM's live brochure URLs for
    every code checked. Extracts the `\\d{2}-\\d{3,4}` code via regex
    rather than a fixed "RI " prefix-strip: at least one row in OPM's
    own spreadsheet has a malformed value ("RI-73 899" instead of
    "RI 73-899", found during this step's validation) that a plain
    prefix-strip would turn into a garbage filename/URL."""
    try:
        content = await _fetch(client, _OPM_PLAN_KEY_URL)
    except httpx.HTTPError as exc:
        logger.warning(
            "opm_plan_key_fetch_failed",
            url=_OPM_PLAN_KEY_URL,
            error=str(exc),
            hint="The plan-key filename embeds the plan year and publish date -- "
            "it likely changed. Update _OPM_PLAN_KEY_URL.",
        )
        return []

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    brochure_col = header.index("Brochure Number")

    codes: set[str] = set()
    for row in rows[1:]:
        raw = row[brochure_col]
        if not raw:
            continue
        match = _BROCHURE_CODE_RE.search(str(raw))
        if match is None:
            logger.warning("unrecognized_brochure_number", raw_value=str(raw))
            continue
        codes.add(match.group(1))

    return [
        DownloadTarget(
            source="opm",
            subfolder="brochures",
            filename=f"{code}.pdf",
            url=_OPM_BROCHURE_URL_TEMPLATE.format(code=code),
        )
        for code in sorted(codes)
    ]


def _curated_targets(source: _SOURCES, docs: list[tuple[str, str, str]]) -> list[DownloadTarget]:
    return [
        DownloadTarget(source=source, subfolder=subfolder, filename=filename, url=url)
        for subfolder, filename, url in docs
    ]


async def _download_all(
    targets: list[DownloadTarget], *, dry_run: bool, force: bool, request_delay_seconds: float
) -> tuple[int, int, int]:
    """Returns (downloaded, skipped_existing, failed)."""
    manifest = Manifest.load()
    downloaded = skipped = failed = 0

    async with httpx.AsyncClient(headers={"User-Agent": _USER_AGENT}) as client:
        for target in targets:
            destination = _OUTPUT_ROOT / target.source / target.subfolder / target.filename
            # `.as_posix()`, not `str()` -- keeps the manifest's keys
            # forward-slashed regardless of host OS.
            relative_path = destination.relative_to(_OUTPUT_ROOT).as_posix()

            if not force and destination.exists() and manifest.has(relative_path):
                skipped += 1
                continue

            if dry_run:
                logger.info("would_download", path=relative_path, url=target.url)
                downloaded += 1
                continue

            try:
                content = await _fetch(client, target.url)
            except httpx.HTTPError as exc:
                logger.warning("download_failed", url=target.url, error=str(exc))
                failed += 1
                continue

            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(content)
            manifest.record(relative_path, url=target.url, content=content)
            manifest.save()
            downloaded += 1
            logger.info("downloaded", path=relative_path, bytes=len(content))

            # Polite pacing -- these are government servers, not a CDN
            # built for bulk scraping; no rate-limit is documented for
            # any of the four sources, so this is a conservative default
            # rather than a measured-safe value. `asyncio.sleep`, not
            # `time.sleep`: this is an async function, and a blocking
            # sleep would stall the event loop for no reason.
            await asyncio.sleep(request_delay_seconds)

    return downloaded, skipped, failed


async def _main(args: argparse.Namespace) -> int:
    async with httpx.AsyncClient(headers={"User-Agent": _USER_AGENT}) as discovery_client:
        targets: list[DownloadTarget] = []
        if args.source in ("opm", "all"):
            targets += await _discover_opm_brochures(discovery_client)
        if args.source in ("cms", "all"):
            targets += _curated_targets("cms", _CMS_DOCS)
        if args.source in ("dol", "all"):
            targets += _curated_targets("dol", _DOL_DOCS)

    if args.limit is not None:
        by_source: dict[str, list[DownloadTarget]] = {}
        for target in targets:
            by_source.setdefault(target.source, []).append(target)
        targets = [t for group in by_source.values() for t in group[: args.limit]]

    logger.info("discovery_complete", target_count=len(targets))
    downloaded, skipped, failed = await _download_all(
        targets, dry_run=args.dry_run, force=args.force, request_delay_seconds=args.delay
    )
    logger.info("done", downloaded=downloaded, skipped=skipped, failed=failed)
    return 1 if failed and downloaded == 0 else 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="List targets without downloading.")
    parser.add_argument(
        "--force", action="store_true", help="Re-download files already in the manifest."
    )
    parser.add_argument(
        "--source",
        choices=["opm", "cms", "dol", "all"],
        default="all",
        help="Restrict to one source.",
    )
    parser.add_argument("--limit", type=int, default=None, help="Cap files downloaded per source.")
    parser.add_argument(
        "--delay", type=float, default=0.5, help="Seconds to sleep between downloads (default 0.5)."
    )
    return parser.parse_args()


if __name__ == "__main__":
    sys.exit(asyncio.run(_main(_parse_args())))
